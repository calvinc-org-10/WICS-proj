import datetime
import os, uuid
from django import forms
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models.query import QuerySet
from django.http import HttpResponseRedirect, HttpResponse, HttpRequest
from django.urls import reverse
from django.shortcuts import render
from django.views.generic import ListView
from typing import *
from cMenu.models import getcParm
from cMenu.utils import makebool, isDate, WrapInQuotes, calvindate
from mathematical_expressions_parser.eval import evaluate
from openpyxl import load_workbook
from userprofiles.models import WICSuser
from WICS.forms import CountEntryForm, RelatedMaterialInfo, RelatedScheduleInfo
from WICS.models import ActualCounts, MaterialList, CountSchedule, WhsePartTypes
from WICS.procs_SAP import fnSAPList



@login_required
def fnCountEntryForm(req, recNum = 0, 
    loadMatlInfo = None, 
    loadCountDate=str(calvindate().today()), 
    gotoCommand=None
    ):
# one day I will clean this up ...

    _userorg = WICSuser.objects.get(user=req.user).org
    FormMain = CountEntryForm
    FormSubs = [S for S in [RelatedMaterialInfo, RelatedScheduleInfo]]

    modelMain = FormMain.Meta.model
    modelSubs = [S.Meta.model for S in FormSubs]

    # the string 'None' is not the same as the value None
    if loadMatlInfo=='None': loadMatlInfo=None
    if gotoCommand=='None': gotoCommand=None

    prefixvals = {}
    prefixvals['main'] = 'counts'
    prefixvals['matl'] = 'matl'
    prefixvals['schedule'] = 'schedule'
    initialvals = {}
    initialvals['main'] = {'CountDate':calvindate(loadCountDate).as_datetime()}
    initialvals['matl'] = {}
    initialvals['schedule'] = {'CountDate':calvindate(loadCountDate).as_datetime()}

    # recover currRec and matlSubFm from POST data
    if req.method == 'POST':
        try:
            recNum = int(req.POST['RecPK'])
        except:
            recNum = 0
        loadMatlInfo = MaterialList.objects.get(org=_userorg, pk=req.POST['MatlPK']).Material

    # if a record number was passed in, retrieve it
    # # later, handle record not found (i.e. - invalid recNum passed in)
    if recNum <= 0:
        currRec = ActualCounts.objects.filter(org=_userorg).none()
        if loadMatlInfo:
            matlRec = MaterialList.objects.get(org=_userorg, Material=loadMatlInfo)
        else:
            matlRec = MaterialList.objects.none()
    else:
        currRec = ActualCounts.objects.filter(org=_userorg).get(pk=recNum)
        matlRec = currRec.Material
    # endif
    MaterialID = getattr(matlRec, 'pk', None)
    
    changes_saved = {
        'main': False,
        'matl': False,
        'schedule': False
        }
    chgd_dat = {
        'main':None, 
        'matl': None, 
        'schedule': None
        }

    if req.method == 'POST':
        # changed data is being submitted.  process and save it
        # process mainFm AND subforms.

        gotoCommand = None

        # process main form
        if currRec: mainFm = CountEntryForm(_userorg, req.POST, instance=currRec,  prefix=prefixvals['main'])   # do I need to pass in intial?
        else: mainFm = CountEntryForm(_userorg, req.POST, initial=initialvals['main'],  prefix=prefixvals['main']) 
        matlSubFm = RelatedMaterialInfo(_userorg, MaterialID, req.POST, instance=matlRec, prefix=prefixvals['matl'])
        #schedSet = RelatedScheduleInfo(_userorg, SchedID, req.POST, prefix=prefixvals['schedule'], initial=initialvals['schedule'])

        s = ActualCounts.objects.none()

        # if mainFm.is_valid() and matlSubFm.is_valid() and schedFm.is_valid():
        if mainFm.is_valid() and matlSubFm.is_valid():
            if mainFm.has_changed():
                s = mainFm.save()
                chgd_dat['main'] = mainFm.changed_data
                changes_saved['main'] = s.id
            # material info subform
            if matlSubFm.has_changed():
                matlSubFm.save()
                chgd_dat['matl'] = matlSubFm.changed_data
                changes_saved['matl'] = True
            # count schedule subform
            # if schedSet.has_changed():
            #      schedSet.save()
            #      chgd_dat['schedule'] = schedSet.changed_data
            #      changes_saved['schedule'] = True

            gotoCommand = "New"
        else:
            gotoCommand = "Invalid"

    #endif req.method=='POST'

    # if this is a gotoCommand, get the correct record
    if gotoCommand=="First" or (gotoCommand=="Prev" and recNum <=0):
        currRec = ActualCounts.objects.filter(org=_userorg).order_by('id').first()
        if currRec: 
            recNum = currRec.id
            loadMatlInfo = currRec.Material.Material
            matlRec = currRec.Material
            MaterialID = matlRec.id
        else: recNum = 0
    elif gotoCommand=="Prev":
        currRec = ActualCounts.objects.filter(org=_userorg,pk__lt=recNum).order_by('id').last()
        if currRec: 
            recNum = currRec.id
            loadMatlInfo = currRec.Material.Material
            matlRec = currRec.Material
            MaterialID = matlRec.id
        else: recNum = 0
    elif gotoCommand=="Next":
        currRec = ActualCounts.objects.filter(org=_userorg,pk__gt=recNum).order_by('id').first()
        if currRec: 
            recNum = currRec.id
            loadMatlInfo = currRec.Material.Material
            matlRec = currRec.Material
            MaterialID = matlRec.id
        else: recNum = 0
    elif gotoCommand=="Last":
        currRec = ActualCounts.objects.filter(org=_userorg).order_by('id').last()
        if currRec: 
            recNum = currRec.id
            loadMatlInfo = currRec.Material.Material
            matlRec = currRec.Material
            MaterialID = matlRec.id
        else: recNum = 0
    elif gotoCommand=="Invalid":
        # this command occurs when a form (new or existing) is submitted, but it has errors
        # currRec is already set above
        if currRec: 
            recNum = currRec.id
            loadMatlInfo = currRec.Material.Material
            matlRec = currRec.Material
            MaterialID = matlRec.id
        else: recNum = 0
    elif gotoCommand=="New":
        currRec = ActualCounts.objects.filter(org=_userorg).none()
        recNum=0
        loadMatlInfo = None
        matlRec = getattr(currRec,'Material', '')
        MaterialID = getattr(matlRec, 'pk', None)
    else:
        if currRec:
            recNum = currRec.id
            loadMatlInfo = currRec.Material.Material
            matlRec = currRec.Material
            MaterialID = matlRec.id
        else:
            # recNum remains whatever was passed in (0 - else there would be a currRec)
            # passedMatlNum remains whatever was passed in
            # matlRec remains what was constructed above
            MaterialID = getattr(matlRec, 'pk', None)
    #endif

    # prep the forms for the template
    # mainFm and matlSubFm
    if gotoCommand != "Invalid":
        if currRec: 
            mainFm = CountEntryForm(_userorg, instance=currRec, prefix=prefixvals['main'])
        else:       
            mainFm = CountEntryForm(_userorg, initial=initialvals['main'],  prefix=prefixvals['main'])
        if matlRec:
            matlSubFm = RelatedMaterialInfo(_userorg, MaterialID, instance=matlRec, prefix=prefixvals['matl'])
        else:
            matlSubFm = RelatedMaterialInfo(_userorg, MaterialID, initial=initialvals['matl'], prefix=prefixvals['matl'])
    # all counts for this Material today
    if matlRec: 
        todayscounts = ActualCounts.objects.filter(CountDate=loadCountDate,Material=matlRec)
    else: 
        todayscounts = ActualCounts.objects.none()
    # schedFm
    if currRec:
        getDate = currRec.CountDate
        if CountSchedule.objects.filter(org=_userorg, CountDate=getDate, Material=matlRec).exists():
            schedinfo = CountSchedule.objects.filter(org=_userorg, CountDate=getDate, Material=matlRec)[0]  # filter rather than get, since a scheduled count may not exist, or multiple may exist (shouldn't but ...)
        else:
            schedinfo = CountSchedule.objects.none()
    elif (loadMatlInfo!=None) and (gotoCommand==None):
        # review and clean up this block!
        if loadMatlInfo != None:
            # fill in MatlInfo and CountSchedInfo
            if recNum > 0: getDate = currRec.CountDate 
            else: getDate = loadCountDate
            if CountSchedule.objects.filter(org=_userorg, CountDate=getDate, Material=matlRec).exists():
                schedinfo = CountSchedule.objects.filter(org=_userorg, CountDate=getDate, Material=matlRec)[0]  # filter rather than get, since a scheduled count may not exist, or multiple may exist (shouldn't but ...)
            else:
                schedinfo = CountSchedule.objects.none()
        elif recNum > 0:
            # ??????????? shouldn't this already be handled?  Think about it...
            # fill in MatlInfo and CountSchedInfo
            getDate = currRec.CountDate
            if CountSchedule.objects.filter(org=_userorg, CountDate=getDate, Material=matlRec).exists():
                schedinfo = CountSchedule.objects.filter(org=_userorg, CountDate=getDate, Material=matlRec)[0]  # filter rather than get, since a scheduled count may not exist, or multiple may exist (shouldn't but ...)
            else:
                schedinfo = CountSchedule.objects.none()
    else: schedinfo = CountSchedule.objects.none()
    if not schedinfo: schedFm = RelatedScheduleInfo(_userorg, None, initial=initialvals['schedule'], prefix=prefixvals['schedule'])
    else: schedFm = RelatedScheduleInfo(_userorg, schedinfo.pk, instance=schedinfo, prefix=prefixvals['schedule'])

    # CountEntryForm MaterialList dropdown
    matlchoiceForm = {}
    if currRec:
        matlchoiceForm['gotoItem'] = currRec
    else:
        if loadMatlInfo==None: loadMatlInfo = ''
        matlchoiceForm['gotoItem'] = {'Material':loadMatlInfo}
    matlchoiceForm['choicelist'] = MaterialList.objects.filter(org=_userorg).values('id','Material')

    # display the form
    cntext = {'frmMain': mainFm,
            'frmMatlInfo': matlSubFm,
            'todayscounts': todayscounts,
            'matlchoiceForm':matlchoiceForm,
            'noSchedInfo':(not schedinfo),
            'frmSchedInfo': schedFm,
            'changes_saved': changes_saved,
            'changed_data': chgd_dat,
            'recNum': recNum,
            'matlnum_changed': loadMatlInfo,
            'orgname':_userorg.orgname, 'uname':req.user.get_full_name()
            }
    templt = 'frm_CountEntry.html'
    return render(req, templt, cntext)

##############################################################
##############################################################
##############################################################

@login_required
def fnUploadActCountSprsht(req):
    _userorg = WICSuser.objects.get(user=req.user).org

    def validatefld(fld, val):
        if   fld == 'CountDate': 
            if isinstance(val,(calvindate, datetime.date, datetime.datetime)):
                retval = True
            else:
                retval = (isDate(val) != False)
        elif fld == 'Material': 
            retval = True
        elif fld == 'Counter': 
            retval = True
        elif fld == 'BLDG': 
            retval = True
        elif fld == 'LOCATION': 
            retval = True
        elif fld == 'LocationOnly': 
            if isinstance(val,str):
                retval = str.isnumeric()
            elif isinstance(val,(float,int)):
                retval = True
            else:                
                retval = False
        elif fld == 'CTD_QTY_Expr': 
            try:
                v = evaluate(val)
            except (SyntaxError, NameError, TypeError, ZeroDivisionError):
                v = "-- INVALID --"
            retval = (v!="-- INVALID --")
        elif fld == 'Notes': 
            retval = True
        elif fld == 'TypicalContainerQty' \
        or fld == 'TypicalPalletQty':
            if val == '' or val == None: 
                retval = True   # this will be converted to 0
            elif isinstance(val,str):
                retval = str.isnumeric()
            elif isinstance(val,(float,int)):
                retval = True
            else:                
                retval = False
        else:
            retval = True
        
        return retval

    if req.method == 'POST':
        # save the file so we can open it as an excel file
        CountSprshtFile = req.FILES['CEFile']
        svdir = getcParm('SAP-FILELOC')
        fName = svdir+"tmpCE"+str(uuid.uuid4())+".xlsx"
        with open(fName, "wb") as destination:
            for chunk in CountSprshtFile.chunks():
                destination.write(chunk)

        wb = load_workbook(filename=fName, read_only=True)
        ws = wb['Counts']

        CountSprshtcolmnNames = ws[1]
        CountSprshtcolmnMap = {'Material': None,'CountDate':None, 'Counter':None, 'BLDG':None}
        CountSprsht_SSName_TableName_map = {
                'CountDate': 'CountDate',
                'Counter': 'Counter',
                'BLDG': 'BLDG',
                'LOCATION': 'LOCATION',
                'Material': 'Material',
                'LocationOnly': 'LocationOnly',
                'CTD_QTY_Expr': 'CTD_QTY_Expr',
                'TypicalContainerQty': 'TypicalContainerQty',
                'TypicalPalletQty': 'TypicalPalletQty',
                'Notes': 'Notes',
                }
        for col in CountSprshtcolmnNames:
            if col.value in CountSprsht_SSName_TableName_map:
                CountSprshtcolmnMap[CountSprsht_SSName_TableName_map[col.value]] = col.column - 1
        if (CountSprshtcolmnMap['Material'] == None) \
          or (CountSprshtcolmnMap['CountDate'] == None) \
          or (CountSprshtcolmnMap['Counter'] == None) \
          or (CountSprshtcolmnMap['BLDG'] == None):
            raise Exception('SAP Spreadsheet has bad header row.  See Calvin to fix this.')

        UplResults = []
        nRows = 0
        rowNum=1
        MAX_COUNT_ROWS = 5000
        for row in ws.iter_rows(min_row=rowNum+1, max_row=MAX_COUNT_ROWS, values_only=True):
            rowNum += 1
            try:
                MatObj = MaterialList.objects.get(org=_userorg, Material=row[CountSprshtcolmnMap['Material']])
            except:
                MatObj = None

            if MatObj:
                requiredFields={
                        'CountDate': False,
                        'Material': False,
                        'Counter': False,
                        'BLDG': False,
                        'Both LocationOnly and CTD_QTY': False,
                        }
                MatChanged = False
                SRec = ActualCounts(org = _userorg)
                for fldName, colNum in CountSprshtcolmnMap.items():
                    V = row[colNum]
                    if V!=None: 
                        if validatefld(fldName, V):
                            if   fldName == 'CountDate': 
                                setattr(SRec, fldName, calvindate(V).as_datetime())   #calvindate
                                requiredFields['CountDate'] = True
                            elif fldName == 'Material': 
                                setattr(SRec, fldName, MatObj)
                                requiredFields['Material'] = True
                            elif fldName == 'Counter': 
                                setattr(SRec, fldName, V)
                                requiredFields['Counter'] = True
                            elif fldName == 'BLDG': 
                                setattr(SRec, fldName, V)
                                requiredFields['BLDG'] = True
                            elif fldName == 'LOCATION': setattr(SRec, fldName, V)
                            elif fldName == 'LocationOnly': 
                                setattr(SRec, fldName, makebool(V))
                                requiredFields['Both LocationOnly and CTD_QTY'] = True
                            elif fldName == 'CTD_QTY_Expr': 
                                setattr(SRec, fldName, V)
                                requiredFields['Both LocationOnly and CTD_QTY'] = True
                            elif fldName == 'Notes': setattr(SRec, fldName, V)
                            elif fldName == 'TypicalContainerQty' \
                            or fldName == 'TypicalPalletQty':
                                if V == '' or V == None: V = 0
                                if V != 0 and V != getattr(MatObj,fldName,0): 
                                    setattr(MatObj, fldName, V)
                                    MatChanged = True
                        else:
                            UplResults.append({'error':str(V)+' is invalid for '+fldName, 'rowNum':rowNum})
                            

                # are all required fields present?
                AllRequiredPresent = True
                for keyname, Prsnt in requiredFields.items():
                    AllRequiredPresent = AllRequiredPresent and Prsnt
                    if not Prsnt:
                        UplResults.append({'error':keyname+' missing', 'rowNum':rowNum})

                if AllRequiredPresent:
                    SRec.save()
                    if MatChanged: MatObj.save()
                    qs = type(SRec).objects.filter(pk=SRec.pk).values().first()
                    res = {'error': False, 'rowNum':rowNum, 'TypicalQty':MatChanged, 'MaterialNum': row[CountSprshtcolmnMap['Material']] }
                    res.update(qs)
                    UplResults.append(res)
                    nRows += 1
            else:
                if row[CountSprshtcolmnMap['Material']]:
                    UplResults.append({'error':row[CountSprshtcolmnMap['Material']]+' does not exist in MaterialList', 'rowNum':rowNum})

        if rowNum >= MAX_COUNT_ROWS:
            UplResults.insert(0,{'error':f'Data in spreadsheet rows {MAX_COUNT_ROWS+1} and beyond are being ignored.'})

        # close and kill temp files
        wb.close()
        os.remove(fName)

        cntext = {'UplResults':UplResults, 'nRowsRead':rowNum, 'nRowsAdded':nRows,
                'orgname':_userorg.orgname, 'uname':req.user.get_full_name()
                }
        templt = 'frm_uploadCountEntry_Success.html'
    else:
        cntext = {'orgname':_userorg.orgname, 'uname':req.user.get_full_name()
                }
        templt = 'frm_UploadCountEntrySprdsht.html'
    #endif

    return render(req, templt, cntext)

#####################################################################
#####################################################################
#####################################################################

class ActualCountListForm(LoginRequiredMixin, ListView):
    ordering = ['-CountDate', 'Material']
    context_object_name = 'ActCtList'
    template_name = 'frm_ActualCountList.html'
    
    def setup(self, req: HttpRequest, *args: Any, **kwargs: Any) -> None:
        self._user = req.user
        self._userorg = WICSuser.objects.get(user=req.user).org
        self.queryset = ActualCounts.objects.filter(org=self._userorg).order_by('-CountDate', 'Material')   # figure out how to pass in self.ordering
        return super().setup(req, *args, **kwargs)

    def get_queryset(self) -> QuerySet[Any]:
        return super().get_queryset()

    # def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return super().get(request, *args, **kwargs)

    # def post(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
    #     return HttpResponse('Not needed!!')

    def render_to_response(self, context: Dict[str, Any], **response_kwargs: Any) -> HttpResponse:
        context.update({'orgname': self._userorg.orgname,  'uname':self._user.get_full_name()})
        return super().render_to_response(context, **response_kwargs)

#####################################################################
#####################################################################
#####################################################################

@login_required
def fnCountSummaryRpt (req, passedCountDate='CURRENT_DATE'):
    _userorg = WICSuser.objects.get(user=req.user).org

    # get the SAP data
    dtobj_pDate = isDate(passedCountDate)
    if not dtobj_pDate: dtobj_pDate = calvindate().as_datetime()
    SAP_SOH = fnSAPList(_userorg,dtobj_pDate)
    
    def CreateOutputRows(raw_qs, Eval_CTDQTY=True):
        def SummaryLine(lastrow):
            # summarize last Matl
            # total SAP Numbers
            SAPTot = 0
            outputline = dict()
            outputline['type'] = 'Summary'
            outputline['SAPNum'] = []
            for SAProw in SAP_SOH['SAPTable'].filter(Material=lastrow['Material']): 
                outputline['SAPNum'].append((SAProw.StorageLocation, SAProw.Amount, SAProw.BaseUnitofMeasure))
                SAPTot += SAProw.Amount*SAProw.mult
            outputline['TypicalContainerQty'] = lastrow['TypicalContainerQty']
            outputline['TypicalPalletQty'] = lastrow['TypicalPalletQty']
            outputline['Material'] = lastrow['Material']
            outputline['PartType'] = lastrow['PartType']
            outputline['CountTotal'] = lastrow['TotalCounted']
            outputline['SAPTotal'] = SAPTot
            outputline['Diff'] = lastrow['TotalCounted'] - SAPTot
            divsr = 1
            if lastrow['TotalCounted']!=0 or SAPTot!=0: divsr = max(lastrow['TotalCounted'], SAPTot)
            outputline['Accuracy'] = min(lastrow['TotalCounted'], SAPTot) / divsr * 100
            outputline['CMFlag'] = lastrow['CMFlag']
            outputline['ReasonScheduled'] = lastrow['ReasonScheduled']
            outputline['SchedNotes'] = lastrow['SchedNotes']
            outputline['MatlNotes'] = lastrow['MatlNotes']
            #outputrows.append(outputline)

            return outputline

        def CreateLastrow(rawrow):
            lastrow = dict()
            lastrow['Material'] = rawrow.Matl_PartNum
            lastrow['PartType'] = rawrow.PartType
            lastrow['TotalCounted'] = 0
            lastrow['CMFlag'] = rawrow.cs_CMPrintFlag
            lastrow['SchedNotes'] = rawrow.cs_Notes
            lastrow['TypicalContainerQty'] = rawrow.TypicalContainerQty
            lastrow['TypicalPalletQty'] = rawrow.TypicalPalletQty
            lastrow['MatlNotes'] = rawrow.mtl_Notes
            lastrow['ReasonScheduled'] = rawrow.cs_ReasonScheduled

            return lastrow

        def DetailLine(rawrow, Eval_CTDQTY=True):
            outputline = dict()
            outputline['type'] = 'Detail'
            outputline['CycCtID'] = rawrow.ac_CycCtID
            outputline['Material'] = rawrow.Matl_PartNum
            outputline['SchedCounter'] = rawrow.cs_Counter
            outputline['ActCounter'] = rawrow.ac_Counter
            outputline['BLDG'] = rawrow.ac_BLDG
            outputline['LOCATION'] = rawrow.ac_LOCATION
            outputline['PKGID'] = rawrow.ac_PKGID_Desc
            outputline['TAGQTY'] = rawrow.ac_TAGQTY
            outputline['PossNotRec'] = rawrow.FLAG_PossiblyNotRecieved
            outputline['MovDurCt'] = rawrow.FLAG_MovementDuringCount
            outputline['CTD_QTY_Expr'] = rawrow.ac_CTD_QTY_Expr
            if Eval_CTDQTY:
                try:
                    outputline['CTD_QTY_Eval'] = evaluate(rawrow.ac_CTD_QTY_Expr)   # yes, I know the risks - Ill write my own parser later ...
                    # do next line at caller
                    # lastrow['TotalCounted'] += outputline['CTD_QTY_Eval']
                except:
                    # Exception('bad expression:'+rawrow.ac_CTD_QTY_Expr)
                    outputline['CTD_QTY_Eval'] = "????"
            else:
                outputline['CTD_QTY_Eval'] = "----"
            outputline['ActCountNotes'] = rawrow.ac_Notes
            # outputrows.append(outputline)

            return outputline
        
        outputrows = []
        lastrow = {'Material': None}
        for rawrow in raw_qs:
            if rawrow.Matl_PartNum != lastrow['Material']:     # new Matl
                if outputrows:
                    outputrows.append(SummaryLine(lastrow))
                # no else -  if outputrows is empty, this is the first row, so keep going

                # this new material is now the "old" one; save values for when it switches, and we do the above block
                # this whole block becomes
                lastrow = CreateLastrow(rawrow)
            #endif

            # process this row
            outputline = DetailLine(rawrow, Eval_CTDQTY)
            outputrows.append(outputline)
            if isinstance(outputline['CTD_QTY_Eval'],(int,float)): lastrow['TotalCounted'] += outputline['CTD_QTY_Eval']
        # endfor
        # need to do the summary on the last row
        if outputrows:
            # summarize last Matl
            outputrows.append(SummaryLine(lastrow))
        
        return outputrows
        
    ### main body of fnCountSummaryRpt

    fldlist = "0 as id, cs.id as cs_id, cs.CountDate as cs_CountDate , cs.Counter as cs_Counter" \
        ", cs.Priority as cs_Priority, cs.ReasonScheduled as cs_ReasonScheduled, cs.CMPrintFlag as cs_CMPrintFlag" \
        ", cs.Notes as cs_Notes" \
        ", ac.id as ac_id, ac.CountDate as ac_CountDate, ac.CycCtID as ac_CycCtID, ac.Counter as ac_Counter" \
        ", ac.LocationOnly as ac_LocationOnly, ac.CTD_QTY_Expr as ac_CTD_QTY_Expr, ac.BLDG as ac_BLDG" \
        ", ac.LOCATION as ac_LOCATION, ac.PKGID_Desc as ac_PKGID_Desc, ac.TAGQTY as ac_TAGQTY" \
        ", ac.FLAG_PossiblyNotRecieved, ac.FLAG_MovementDuringCount, ac.Notes as ac_Notes" \
        ", mtl.Material as Matl_PartNum, (SELECT WhsePartType FROM WICS_whseparttypes WHERE id=mtl.PartType_id) as PartType" \
        ", mtl.Description, mtl.TypicalContainerQty, mtl.TypicalPalletQty, mtl.Notes as mtl_Notes"
    org_condition = '(ac.org_id = ' + str(_userorg.pk) + ' OR cs.org_id = ' + str(_userorg.pk) + ') '
    if isDate(passedCountDate): datestr = WrapInQuotes(passedCountDate,"'","'")
    else: datestr = passedCountDate
    date_condition = '(ac.CountDate = ' + datestr + ' OR cs.CountDate = ' + datestr + ') '
    order_by = 'Matl_PartNum'

    SummaryReport = []

    A_Sched_Ctd_from = 'WICS_countschedule cs INNER JOIN WICS_materiallist mtl INNER JOIN WICS_actualcounts ac'    
    A_Sched_Ctd_joinon = 'cs.CountDate=ac.CountDate AND cs.Material_id=ac.Material_id AND ac.Material_id=mtl.id'    
    A_Sched_Ctd_where = ''
    A_Sched_Ctd_sql = 'SELECT ' + fldlist + \
        ' FROM ' + A_Sched_Ctd_from + \
        ' ON ' + A_Sched_Ctd_joinon + \
        ' WHERE NOT ac.LocationOnly AND ' + org_condition + ' AND ' + date_condition + \
        ' ORDER BY ' + order_by
        # + ' AND ' + A_Sched_Ctd_where
    A_Sched_Ctd_qs = CountSchedule.objects.raw(A_Sched_Ctd_sql)
    # build display lines
    SummaryReport.append({
                'Title':'Scheduled and Counted',
                'outputrows': CreateOutputRows(A_Sched_Ctd_qs)
                })

    B_UnSched_Ctd_from = 'WICS_countschedule cs RIGHT JOIN' \
        ' (WICS_actualcounts ac INNER JOIN WICS_materiallist mtl ON ac.Material_id=mtl.id)'
    B_UnSched_Ctd_joinon = 'cs.CountDate=ac.CountDate AND cs.Material_id=ac.Material_id'
    B_UnSched_Ctd_where = '(cs.id IS NULL)'
    B_UnSched_Ctd_sql = 'SELECT ' + fldlist + ' ' + \
        ' FROM ' + B_UnSched_Ctd_from + \
        ' ON ' + B_UnSched_Ctd_joinon + \
        ' WHERE NOT ac.LocationOnly AND ' + org_condition + ' AND ' + date_condition + \
        ' AND ' + B_UnSched_Ctd_where + \
        ' ORDER BY ' + order_by
    B_UnSched_Ctd_qs = CountSchedule.objects.raw(B_UnSched_Ctd_sql)
    SummaryReport.append({
                'Title':'UnScheduled',
                'outputrows': CreateOutputRows(B_UnSched_Ctd_qs)
                })
    
    C_Sched_NotCtd_Ctd_from = '(WICS_countschedule cs INNER JOIN WICS_materiallist mtl ON cs.Material_id=mtl.id)' \
        ' LEFT JOIN WICS_actualcounts ac'
    C_Sched_NotCtd_Ctd_joinon = 'cs.CountDate=ac.CountDate AND cs.Material_id=ac.Material_id'
    C_Sched_NotCtd_Ctd_where = '(ac.id IS NULL)'
    C_Sched_NotCtd_Ctd_sql = 'SELECT ' + fldlist + ' ' + \
        ' FROM ' + C_Sched_NotCtd_Ctd_from + \
        ' ON ' + C_Sched_NotCtd_Ctd_joinon + \
        ' WHERE ' + org_condition + ' AND ' + date_condition + \
        ' AND ' + C_Sched_NotCtd_Ctd_where + \
        ' ORDER BY ' + order_by
    C_Sched_NotCtd_Ctd_qs = CountSchedule.objects.raw(C_Sched_NotCtd_Ctd_sql)
    SummaryReport.append({
                'Title':'Scheduled but Not Counted',
                'outputrows': CreateOutputRows(C_Sched_NotCtd_Ctd_qs, Eval_CTDQTY=False)
                })
    AccuracyCutoff = { 
                'DANGER': float(getcParm('ACCURACY-DANGER')),
                'SUCCESS': float(getcParm('ACCURACY-SUCCESS')),
                'WARNING': float(getcParm('ACCURACY-WARNING')),
                }

    # display the form
    cntext = {
            'CountDate': dtobj_pDate,
            'SAPDate': SAP_SOH['SAPDate'],
            'AccuracyCutoff': AccuracyCutoff,
            'SummaryReport': SummaryReport,
            'orgname':_userorg.orgname, 'uname':req.user.get_full_name()
            }
            #'Sched_Ctd': A_Sched_Ctd_outputrows,
            #'UnSched_Ctd': B_UnSched_Ctd_outputrows,
            #'Sched_NotCtd':C_Sched_NotCtd_Ctd_outputrows,
    templt = 'rpt_CountSummary.html'
    return render(req, templt, cntext)


